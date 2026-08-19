import os
import shutil

from fastapi import (
    APIRouter,
    UploadFile,
    File,
    Depends,
    HTTPException,
)

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from database import get_db
from models import VerifiedFarmer


router = APIRouter(
    prefix="/admin",
    tags=["Admin"],
)


UPLOAD_FOLDER = "uploads/admin"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@router.post("/import-farmers")
async def import_verified_farmers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):

    # Check file type
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Please upload an Excel (.xlsx) file.",
        )

    temp_file = os.path.join(
        UPLOAD_FOLDER,
        "temp_farmers.xlsx",
    )

    try:

        # Save uploaded Excel file
        with open(temp_file, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Open Excel workbook
        workbook = load_workbook(temp_file)

        sheet = workbook.active

        imported = 0
        skipped = 0

        # Expected Excel columns:
        #
        # Column A = Farmer ID
        # Column B = Phone
        # Column C = Age

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):

            farmer_id = row[0]
            phone = row[1]
            age = row[2]

            # Skip empty rows
            if not farmer_id:
                continue

            # Check if Farmer ID already exists
            existing = (
                db.query(VerifiedFarmer)
                .filter(
                    VerifiedFarmer.farmer_id == str(farmer_id)
                )
                .first()
            )

            if existing:
                skipped += 1
                continue

            farmer = VerifiedFarmer(
                farmer_id=str(farmer_id),
                phone=str(phone) if phone else None,
                age=int(age) if age is not None else None,
            )

            db.add(farmer)

            imported += 1

        db.commit()

        return {
            "success": True,
            "message": "Verified farmers imported successfully.",
            "farmers_imported": imported,
            "farmers_skipped": skipped,
        }

    except Exception as e:

        db.rollback()

        raise HTTPException(
            status_code=500,
            detail=f"Error importing farmers: {str(e)}",
        )

    finally:

        if os.path.exists(temp_file):
            os.remove(temp_file)